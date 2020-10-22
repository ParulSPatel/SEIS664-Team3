from sprint import Sprint
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def get_sprints(min_row=3, max_row=22, min_col=6, max_col=6):
    workbook = load_workbook(filename="CostofDelayPrioritizationCalculator.xlsx")
    sheet = workbook['Calculations']
    return int(max(sheet.iter_rows(min_row=min_row,
                                   max_row=max_row,
                                   min_col=min_col,
                                   max_col=max_col,
                                   values_only=True))[0])


def write_sprint_daily_results(sprints, start_column=9):
    '''
    I'll just apologize now for what's going on below.
    '''
    print("Writing daily results")
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    days_row = 4
    days_column = start_column
    day_counter = 1
    earnings_row = 5
    earnings_column = start_column
    header_row = 3
    final_daily_dollars = 0
    header_column = start_column
    workbook = load_workbook(filename="CostofDelayPrioritizationCalculator.xlsx")
    sheet = workbook['Calculations']
    for sprint in sprints:
        cell = sheet.cell(row=header_row, column=header_column, value=f"Sprint {sprint.sprint_num}")
        cell.font = bold_font
        cell.alignment = center_aligned_text
        sheet.merge_cells(start_row=header_row, start_column=header_column, end_row=header_row, end_column=header_column+4)
        for idx, (_, dollars) in enumerate(sprint.earnings.items(), start=1):
            sheet.cell(row=days_row, column=days_column, value=day_counter).fill = PatternFill(start_color="D3D3D3", fill_type="solid")
            sheet.cell(row=earnings_row, column=earnings_column, value=dollars)
            final_daily_dollars = dollars
            day_counter += 1
            days_column += 1
            earnings_column += 1
            if idx % 5 == 0:
                days_row += 2
                days_column = start_column
                earnings_row += 2
                earnings_column = start_column
        header_row = earnings_row + 1
        earnings_column = start_column
        days_column = start_column
        days_row = header_row + 1
        earnings_row = header_row + 2
    workbook.save("CostofDelayPrioritizationCalculator.xlsx")
    return day_counter, final_daily_dollars


def write_totals(sprints, revenue_cell='E37', earnings_cell='F37', staff_pay_cell='D37'):
    print("writing totals")
    total_earnings = 0
    total_revenue = 0
    staff_pay = 0
    # get sprint tallies
    for sprint in sprints:
        total_revenue += sprint.total_earn
        total_earnings += sprint.profit
        staff_pay += sprint.team_pay
    workbook = load_workbook(filename="CostofDelayPrioritizationCalculator.xlsx")
    sheet = workbook['Calculations']
    sheet[revenue_cell] = total_revenue
    sheet[earnings_cell] = total_earnings
    sheet[staff_pay_cell] = staff_pay
    workbook.save("CostofDelayPrioritizationCalculator.xlsx")


def build_sprints(max_sprints,
                  load_min_row,
                  load_max_row,
                  load_min_col,
                  load_max_col):
    sprints = []
    daily_earn = 0

    # create and run sprints, passing the ending daily earn from each
    # into the next so that daily earn starts where the last sprint left off
    for i in range(max_sprints):
        # tell the last sprint in the list that it can stop before 40 days
        last = True if i + 1 == max_sprints else False
        sprint = Sprint(i+1, daily_earn, last)
        sprint.load(load_min_row, load_max_row, load_min_col, load_max_col)
        sprint.run()
        sprints.append(sprint)
        daily_earn += sprint.daily_earn
    return sprints


def main():
    '''
    Loops are for other people
    '''
    max_sprints = get_sprints()
    print(max_sprints)
    sprints = build_sprints(max_sprints, 3, 22, 1, 7)
    write_totals(sprints)
    num_days_iter_one, iter_one_final_daily = write_sprint_daily_results(sprints)

    max_sprints = get_sprints(42, 61, 6, 6)
    print(max_sprints)
    sprints = build_sprints(max_sprints, 42, 61, 1, 7)
    write_totals(sprints, 'E75', 'F75', 'D75')
    num_days_iter_two, iter_two_final_daily = write_sprint_daily_results(sprints, 15)

    # if the time-efficient iteration is shorter than the WSJF iteration we can show
    # how the second iteration would have continued to accrue income in the time
    # that the WSJF iteration was still running and showing revenue in our calculations
    iter_diff = num_days_iter_one - num_days_iter_two
    if iter_diff > 0:
        workbook = load_workbook(filename="CostofDelayPrioritizationCalculator.xlsx")
        sheet = workbook['Calculations']
        sheet['C78'] = iter_two_final_daily * iter_diff
        workbook.save('CostofDelayPrioritizationCalculator.xlsx')


main()
