from openpyxl import load_workbook
# from pprint import PrettyPrinter
from collections import OrderedDict
from feature import Feature


class Sprint():
    def __init__(self, sprint_num, daily_earn, last_sprint=False):
        self.sprint_num = int(sprint_num)
        self.length = 40
        self.working_days = 0
        self.earnings = OrderedDict()
        self.daily_earn = daily_earn
        self.team_pay = 20000
        self.current_day = 1
        self.features = []
        self.total_earn = 0
        self.profit = 0
        self.last_sprint = last_sprint

    def run(self):
        for feature in self.features:
            print(f"operating on feature order: {feature.order}")
            self.earnings, self.current_day, self.daily_earn = feature.run(self.earnings, self.current_day, self.daily_earn)
        # finish out the sprint if features don't use the entire time
        last_worked_day = self.working_days
        if self.working_days < self.length and not self.last_sprint:
            for i in range(self.length-self.working_days):
                last_worked_day += 1
                self.earnings[last_worked_day] = self.daily_earn
        self.total_earn = sum(list(self.earnings.values()))
        self.profit = self.total_earn - self.team_pay

    def load(self, min_row=3, max_row=22, min_col=1, max_col=7):
        workbook = load_workbook(filename="CostofDelayPrioritizationCalculator.xlsx", data_only=True, read_only=True)
        sheet = workbook['Calculations']
        for value in sheet.iter_rows(min_row=min_row,
                                     max_row=max_row,
                                     min_col=min_col,
                                     max_col=max_col,
                                     values_only=True):
            if value[5] == self.sprint_num:
                feature = Feature(value[0], value[1], value[6], value[2])
                self.working_days += feature.days
                self.features.append(feature)
